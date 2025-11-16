import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import time
import os
import gc
import logging
import sys
import sqlite3
import subprocess
import json
from threading import Lock, Thread
import signal
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import queue
import re
import argparse

# 修复Python 3.12中SQLite datetime适配器的弃用警告
def adapt_datetime(dt):
    return dt.isoformat()

def convert_datetime(s):
    return datetime.fromisoformat(s.decode())

sqlite3.register_adapter(datetime, adapt_datetime)
sqlite3.register_converter("timestamp", convert_datetime)

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    print("注意: 未安装tqdm库，将使用简单进度指示")

# 配置日志
logging.basicConfig(
    filename='crawler.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='a'
)
logger = logging.getLogger()

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
logger.addHandler(console_handler)

# Cookie配置
COOKIES = {
    "sessionid": "xe9n2pkue11qtkams0alrm9mgpkwwtsr",
    "csrftoken": "1vTuvnNMyjQswtukKd8q7dPZDOW22aEu"
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Android 12; Mobile) Python Script",
    "Referer": "https://m.mugzone.net/"
}

BASE_URL = "https://m.mugzone.net/page/all/player?from=0&mode={mode}"
PLAYER_PROFILE_URL = "https://m.mugzone.net/accounts/user/{player_id}"
MODES = list(range(10))

DB_FILE = "malody_rankings.db"

GIT_REPO_PATH = os.path.dirname(os.path.abspath(__file__))
GIT_COMMIT_MESSAGE = datetime.now().strftime("%Y-%m-%d %H:%M updated")

stop_requested = False
stop_lock = Lock()

# 玩家配置文件
PLAYER_CONFIG_FILE = "players.txt"

# 玩家爬取队列和状态
player_queue = queue.Queue()
player_crawl_lock = Lock()
player_crawl_in_progress = False
last_player_crawl_time = None

def signal_handler(sig, frame):
    global stop_requested
    with stop_lock:
        stop_requested = True
    logger.info("收到终止信号，正在安全退出...")
    time.sleep(1)
    DatabaseManager().close_connection()
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

def get_git_commit_message():
    """生成Git提交消息"""
    return datetime.now().strftime("%Y-%m-%d %H:%M updated")

class DatabaseManager:
    _instance = None
    _lock = Lock()
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
            cls._instance.connection = None
            cls._instance.connections = {}
        return cls._instance
    
    def get_connection(self, thread_id=None):
        if thread_id is None:
            thread_id = threading.get_ident()
            
        with self._lock:
            if thread_id not in self.connections:
                self.connections[thread_id] = sqlite3.connect(
                    DB_FILE, 
                    detect_types=sqlite3.PARSE_DECLTYPES,
                    timeout=30,
                    check_same_thread=False
                )
                self.connections[thread_id].execute("PRAGMA journal_mode=WAL")
                self.connections[thread_id].execute("PRAGMA busy_timeout = 30000")
            return self.connections[thread_id]
    
    def close_connection(self, thread_id=None):
        with self._lock:
            if thread_id is None:
                for conn in self.connections.values():
                    conn.close()
                self.connections = {}
            elif thread_id in self.connections:
                self.connections[thread_id].close()
                del self.connections[thread_id]
    
    def execute_query(self, query, params=None, thread_id=None):
        conn = self.get_connection(thread_id)
        cursor = conn.cursor()
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            conn.commit()
            return cursor
        except Exception as e:
            conn.rollback()
            raise e
    
    def executemany_query(self, query, params_list, thread_id=None):
        conn = self.get_connection(thread_id)
        cursor = conn.cursor()
        try:
            cursor.executemany(query, params_list)
            conn.commit()
            return cursor
        except Exception as e:
            conn.rollback()
            raise e

def migrate_database():
    """迁移数据库，添加uid字段"""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    logger.info("开始数据库迁移...")
    
    try:
        # 检查是否已经迁移过
        cursor.execute("PRAGMA table_info(player_identity)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_identity ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_identity_uid ON player_identity(uid)')
            logger.info("已添加uid字段到player_identity表")
        
        cursor.execute("PRAGMA table_info(player_aliases)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_aliases ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_aliases_uid ON player_aliases(uid)')
            logger.info("已添加uid字段到player_aliases表")
        
        cursor.execute("PRAGMA table_info(player_rankings)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'uid' not in columns:
            cursor.execute('ALTER TABLE player_rankings ADD COLUMN uid TEXT')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_rankings_uid ON player_rankings(uid)')
            logger.info("已添加uid字段到player_rankings表")
        
        db_manager.get_connection().commit()
        
        with open('sql_changes.md', 'w', encoding='utf-8') as f:
            f.write("# SQL数据库结构变更记录\n\n")
            f.write("## 版本 2.0 - 添加UID支持\n\n")
            f.write("### 变更内容\n\n")
            f.write("1. 在 `player_identity` 表中添加 `uid` 字段\n")
            f.write("2. 在 `player_aliases` 表中添加 `uid` 字段\n")
            f.write("3. 在 `player_rankings` 表中添加 `uid` 字段\n")
            f.write("4. 为各表的 `uid` 字段创建索引\n\n")
            f.write("### SQL语句\n\n")
            f.write("```sql\n")
            f.write("-- 添加uid字段\n")
            f.write("ALTER TABLE player_identity ADD COLUMN uid TEXT;\n")
            f.write("ALTER TABLE player_aliases ADD COLUMN uid TEXT;\n")
            f.write("ALTER TABLE player_rankings ADD COLUMN uid TEXT;\n\n")
            f.write("-- 创建索引\n")
            f.write("CREATE INDEX idx_player_identity_uid ON player_identity(uid);\n")
            f.write("CREATE INDEX idx_player_aliases_uid ON player_aliases(uid);\n")
            f.write("CREATE INDEX idx_player_rankings_uid ON player_rankings(uid);\n")
            f.write("```\n")
        
        logger.info("数据库迁移完成，变更已记录到 sql_changes.md")
        return True
    except Exception as e:
        logger.error("数据库迁移失败: %s", e)
        db_manager.get_connection().rollback()
        return False

def init_database():
    """初始化数据库，创建表结构"""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    try:
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_identity (
            player_id INTEGER PRIMARY KEY AUTOINCREMENT,
            uid TEXT,
            current_name TEXT NOT NULL,
            first_seen TIMESTAMP NOT NULL,
            last_seen TIMESTAMP NOT NULL
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_aliases (
            alias_id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER NOT NULL,
            uid TEXT,
            alias TEXT NOT NULL,
            first_seen TIMESTAMP NOT NULL,
            last_seen TIMESTAMP NOT NULL,
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id)
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_rankings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER,
            uid TEXT,
            mode INTEGER NOT NULL,
            rank INTEGER NOT NULL,
            name TEXT NOT NULL,
            lv INTEGER,
            exp INTEGER,
            acc REAL,
            combo INTEGER,
            pc INTEGER,
            crawl_time TIMESTAMP NOT NULL,
            FOREIGN KEY (player_id) REFERENCES player_identity (player_id),
            UNIQUE(mode, rank, crawl_time)
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS import_metadata (
            mode INTEGER PRIMARY KEY,
            last_import_time TIMESTAMP
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_config (
            config_id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_identifier TEXT NOT NULL UNIQUE,
            player_name TEXT,
            is_active BOOLEAN DEFAULT 1,
            priority INTEGER DEFAULT 5,
            notes TEXT,
            created_time TIMESTAMP NOT NULL,
            last_updated TIMESTAMP NOT NULL
        )
        ''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS player_crawl_status (
            player_identifier TEXT PRIMARY KEY,
            last_crawled TIMESTAMP,
            crawl_count INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            last_error TEXT
        )
        ''')
        
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_identity_uid ON player_identity(uid)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_aliases_uid ON player_aliases(uid)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_player_rankings_uid ON player_rankings(uid)')
        
        for mode in MODES:
            cursor.execute(
                "INSERT OR IGNORE INTO import_metadata (mode, last_import_time) VALUES (?, NULL)",
                (mode,)
            )
        
        db_manager.get_connection().commit()
        logger.info("数据库初始化完成")
        
        migrate_database()
        
    except Exception as e:
        logger.error("数据库初始化失败: %s", e)
        raise

def resolve_player_identity(name, crawl_time, uid=None):
    """解析玩家身份，优先使用uid，处理改名情况"""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    try:
        player_id = None
        
        if uid:
            try:
                cursor.execute(
                    "SELECT player_id FROM player_identity WHERE uid = ?",
                    (uid,)
                )
                result = cursor.fetchone()
                
                if result:
                    player_id = result[0]
                    
                    cursor.execute(
                        "UPDATE player_identity SET last_seen = ?, current_name = ? WHERE player_id = ?",
                        (crawl_time, name, player_id)
                    )
                    
                    cursor.execute(
                        "SELECT alias_id FROM player_aliases WHERE player_id = ? AND alias = ?",
                        (player_id, name)
                    )
                    if not cursor.fetchone():
                        cursor.execute(
                            "INSERT INTO player_aliases (player_id, uid, alias, first_seen, last_seen) VALUES (?, ?, ?, ?, ?)",
                            (player_id, uid, name, crawl_time, crawl_time)
                        )
                    else:
                        cursor.execute(
                            "UPDATE player_aliases SET last_seen = ? WHERE player_id = ? AND alias = ?",
                            (crawl_time, player_id, name)
                        )
            except sqlite3.OperationalError as e:
                if "no such column: uid" in str(e):
                    logger.warning("uid列不存在，回退到名称查找")
                    pass
                else:
                    raise
        
        if not player_id:
            cursor.execute(
                "SELECT player_id FROM player_aliases WHERE alias = ?",
                (name,)
            )
            result = cursor.fetchone()
            
            if result:
                player_id = result[0]
                
                if uid:
                    try:
                        cursor.execute(
                            "UPDATE player_identity SET uid = ? WHERE player_id = ?",
                            (uid, player_id)
                        )
                        cursor.execute(
                            "UPDATE player_aliases SET uid = ? WHERE player_id = ?",
                            (uid, player_id)
                        )
                    except sqlite3.OperationalError as e:
                        if "no such column: uid" in str(e):
                            logger.warning("uid列不存在，跳过uid更新")
                        else:
                            raise
                
                cursor.execute(
                    "UPDATE player_aliases SET last_seen = ? WHERE alias = ?",
                    (crawl_time, name)
                )
                cursor.execute(
                    "UPDATE player_identity SET last_seen = ?, current_name = ? WHERE player_id = ?",
                    (crawl_time, name, player_id)
                )
            else:
                cursor.execute(
                    "INSERT INTO player_identity (uid, current_name, first_seen, last_seen) VALUES (?, ?, ?, ?)",
                    (uid, name, crawl_time, crawl_time)
                )
                player_id = cursor.lastrowid
                
                cursor.execute(
                    "INSERT INTO player_aliases (player_id, uid, alias, first_seen, last_seen) VALUES (?, ?, ?, ?, ?)",
                    (player_id, uid, name, crawl_time, crawl_time)
                )
        
        db_manager.get_connection().commit()
        return player_id
    except Exception as e:
        logger.error("解析玩家身份失败: %s", e)
        db_manager.get_connection().rollback()
        return None

def link_player_aliases(original_name, new_name, change_time):
    """手动关联玩家的两个名字（处理改名）"""
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    try:
        cursor.execute(
            "SELECT player_id FROM player_aliases WHERE alias = ?",
            (original_name,)
        )
        result = cursor.fetchone()
        
        if not result:
            logger.error("找不到原始名字: %s", original_name)
            return False
        
        player_id = result[0]
        
        cursor.execute(
            "SELECT player_id FROM player_aliases WHERE alias = ?",
            (new_name,)
        )
        result = cursor.fetchone()
        
        if result:
            old_player_id = result[0]
            
            cursor.execute(
                "UPDATE player_rankings SET player_id = ? WHERE player_id = ?",
                (player_id, old_player_id)
            )
            
            cursor.execute(
                "UPDATE player_aliases SET player_id = ? WHERE player_id = ?",
                (player_id, old_player_id)
            )
            
            cursor.execute(
                "DELETE FROM player_identity WHERE player_id = ?",
                (old_player_id,)
            )
        else:
            cursor.execute(
                "INSERT INTO player_aliases (player_id, alias, first_seen, last_seen) VALUES (?, ?, ?, ?)",
                (player_id, new_name, change_time, change_time)
            )
        
        cursor.execute(
            "UPDATE player_identity SET current_name = ? WHERE player_id = ?",
            (new_name, player_id)
        )
        
        db_manager.get_connection().commit()
        logger.info("成功关联玩家改名: %s -> %s", original_name, new_name)
        return True
    except Exception as e:
        logger.error("处理玩家改名失败: %s", e)
        db_manager.get_connection().rollback()
        return False

def parse_player_list(html):
    """解析排行榜页面，返回玩家数据，包括玩家ID"""
    soup = BeautifulSoup(html, "html.parser")
    players = []
    
    top_items = soup.select("div.item-top")
    for item in top_items:
        label_tag = item.select_one("i.label")
        rank = None
        if label_tag and label_tag.has_attr("class"):
            for c in label_tag["class"]:
                if c.startswith("top-"):
                    rank = c.replace("top-", "")
                    break
        
        name_tag = item.select_one("span.name a")
        lv_tag = item.select_one("span.lv")
        acc_tag = item.select_one("span.acc")
        combo_tag = item.select_one("span.combo")
        pc_tag = item.select_one("span.pc, span[class*=pc]")
        
        player_id = None
        if name_tag and name_tag.has_attr('href'):
            href = name_tag['href']
            match = re.search(r'/accounts/user/(\d+)', href)
            if match:
                player_id = match.group(1)
        
        level = None
        exp = None
        if lv_tag:
            lv_text = lv_tag.text.strip()
            if '-' in lv_text:
                parts = lv_text.split('-')
                level = parts[0].replace("Lv.", "").strip()
                exp = parts[1].strip()
            else:
                level = lv_text.replace("Lv.", "").strip()
        
        acc_text = None
        if acc_tag:
            acc_text = acc_tag.text.replace("Acc:", "").replace("%", "").strip()
        
        combo_text = None
        if combo_tag:
            combo_text = combo_tag.text.replace("Combo:", "").strip()
        
        playcount = None
        if pc_tag:
            pc_text = pc_tag.text.replace("游玩次数:", "").strip()
            digits = ''.join(filter(str.isdigit, pc_text))
            if digits:
                playcount = int(digits)
        
        players.append({
            "rank": rank,
            "name": name_tag.text.strip() if name_tag else None,
            "player_id": player_id,
            "lv": level,
            "exp": exp,
            "acc": acc_text,
            "combo": combo_text,
            "pc": playcount
        })

    list_items = soup.select("div.item")
    for item in list_items:
        rank_tag = item.select_one("span.rank")
        rank = rank_tag.text.strip() if rank_tag else None
        
        name_tag = item.select_one("span.name a")
        lv_tag = item.select_one("span.lv")
        exp_tag = item.select_one("span.exp")
        acc_tag = item.select_one("span.acc")
        pc_tag = item.select_one("span.pc, span[class*=pc]")
        combo_tag = item.select_one("span.combo")
        
        player_id = None
        if name_tag and name_tag.has_attr('href'):
            href = name_tag['href']
            match = re.search(r'/accounts/user/(\d+)', href)
            if match:
                player_id = match.group(1)
        
        acc_text = None
        if acc_tag:
            acc_text = acc_tag.text.replace("%", "").strip()
        
        playcount = None
        if pc_tag:
            pc_text = pc_tag.text.strip()
            digits = ''.join(filter(str.isdigit, pc_text))
            if digits:
                playcount = int(digits)
        
        players.append({
            "rank": rank,
            "name": name_tag.text.strip() if name_tag else None,
            "player_id": player_id,
            "lv": lv_tag.text.strip() if lv_tag else None,
            "exp": exp_tag.text.strip() if exp_tag else None,
            "acc": acc_text,
            "combo": combo_tag.text.strip() if combo_tag else None,
            "pc": playcount
        })
    
    processed_players = []
    for p in players:
        try:
            rank = int(p["rank"]) if p["rank"] else None
        except:
            rank = None
            
        try:
            lv = int(p["lv"]) if p["lv"] else 0
        except:
            lv = 0
            
        try:
            exp = int(p["exp"]) if p["exp"] else 0
        except:
            exp = 0
            
        try:
            acc = float(p["acc"]) if p["acc"] else 0.0
        except:
            acc = 0.0
            
        try:
            combo = int(p["combo"]) if p["combo"] else 0
        except:
            combo = 0
            
        pc = p["pc"] if p["pc"] is not None else 0
        
        if rank is not None:
            processed_players.append({
                "rank": rank,
                "name": p["name"],
                "player_id": p["player_id"],
                "lv": lv,
                "exp": exp,
                "acc": acc,
                "combo": combo,
                "pc": pc
            })
    
    return processed_players

def parse_player_profile(html, player_id):
    """解析玩家个人主页数据"""
    soup = BeautifulSoup(html, "html.parser")
    player_data = []
    
    name_tag = soup.select_one("div.user_head .name span")
    player_name = name_tag.text.strip() if name_tag else f"玩家_{player_id}"
    
    rank_items = soup.select("div.rank .item")
    
    for item in rank_items:
        try:
            img_tag = item.select_one("img")
            if not img_tag or not img_tag.has_attr('src'):
                continue
                
            src = img_tag['src']
            mode = None
            if 'mode-0.png' in src:
                mode = 0
            elif 'mode-1.png' in src:
                mode = 1
            elif 'mode-2.png' in src:
                mode = 2
            elif 'mode-3.png' in src:
                mode = 3
            elif 'mode-4.png' in src:
                mode = 4
            elif 'mode-5.png' in src:
                mode = 5
            elif 'mode-6.png' in src:
                mode = 6
            elif 'mode-7.png' in src:
                mode = 7
            elif 'mode-8.png' in src:
                mode = 8
            elif 'mode-9.png' in src:
                mode = 9
            else:
                continue
            
            rank_tag = item.select_one("p.rank")
            if not rank_tag:
                continue
                
            rank_text = rank_tag.text.strip()
            if rank_text.startswith('#'):
                try:
                    rank = int(rank_text[1:].replace(',', ''))
                except:
                    continue
            else:
                continue
            
            data_spans = item.select("p span")
            exp = 0
            playcount = 0
            acc = 0.0
            combo = 0
            
            for span in data_spans:
                text = span.text.strip()
                if text.startswith('Exp.'):
                    try:
                        exp = int(text.replace('Exp.', '').strip().replace(',', ''))
                    except:
                        pass
                elif text.startswith('Playcount:'):
                    try:
                        playcount = int(text.replace('Playcount:', '').strip())
                    except:
                        pass
                elif text.startswith('Acc.'):
                    try:
                        acc = float(text.replace('Acc.', '').replace('%', '').strip())
                    except:
                        pass
                elif text.startswith('Combo:'):
                    try:
                        combo = int(text.replace('Combo:', '').strip())
                    except:
                        pass
            
            player_data.append({
                "rank": rank,
                "name": player_name,
                "lv": 0,
                "exp": exp,
                "acc": acc,
                "combo": combo,
                "pc": playcount,
                "mode": mode
            })
            
        except Exception as e:
            logger.warning("解析玩家 %s 模式 %s 数据时出错: %s", player_id, mode, e)
            continue
    
    return player_data

def crawl_player_profile(session, player_identifier):
    """爬取玩家个人主页数据"""
    try:
        if not player_identifier.isdigit():
            logger.warning("玩家标识符必须是数字ID: %s", player_identifier)
            return None
        
        url = PLAYER_PROFILE_URL.format(player_id=player_identifier)
        
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        
        player_data = parse_player_profile(resp.text, player_identifier)
        return player_data
        
    except requests.exceptions.RequestException as e:
        logger.error("爬取玩家 %s 个人主页失败: %s", player_identifier, e)
        return None
    except Exception as e:
        logger.error("处理玩家 %s 数据时出错: %s", player_identifier, e)
        return None

def get_excel_filename(mode):
    if mode == 0:
        return "key.xlsx"
    elif mode == 3:
        return "catch.xlsx"
    else:
        return f"mode{mode}.xlsx"

def crawl_mode_player(session, mode):
    url = BASE_URL.format(mode=mode)
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error("模式 %d 请求失败: %s", mode, e)
        return pd.DataFrame()
    
    players = parse_player_list(resp.text)
    
    df = pd.DataFrame(players)
    
    if not df.empty:
        df = df[df['rank'].notnull()]
        df['rank'] = df['rank'].astype(int)
        df = df.sort_values('rank').reset_index(drop=True)
    
    return df

def save_data_to_excel(mode, df, timestamp):
    """保存数据到Excel文件（仅在启用Excel保存时使用）"""
    if df.empty:
        logger.warning("模式 %d 无有效数据，跳过保存", mode)
        return
    
    filename = get_excel_filename(mode)
    sheet_name = f"mode_{mode}"

    try:
        if not os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name)

        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name)

        sub_sheets = [s for s in wb.sheetnames if s.startswith(f"{sheet_name}_")]
        latest_sheet = None
        latest_time = None
        for s in sub_sheets:
            try:
                dt_str = s.replace(f"{sheet_name}_", "")
                dt = datetime.strptime(dt_str, "%Y-%m-%d_%H-%M")
                if latest_time is None or dt > latest_time:
                    latest_time = dt
                    latest_sheet = s
            except:
                continue

        if latest_sheet:
            df_prev = pd.read_excel(filename, sheet_name=latest_sheet)
            if not df_prev.empty and df_prev.equals(df):
                logger.info("模式 %d 数据未变化，跳过保存", mode)
                return

        sub_sheet_name = f"{sheet_name}_{timestamp.strftime('%Y-%m-%d_%H-%M')}"
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sub_sheet_name, index=False)

        logger.info("模式 %d 数据保存到 %s -> %s", mode, filename, sub_sheet_name)
    except Exception as e:
        logger.exception("保存模式 %d 数据到Excel失败", mode)

def save_to_database(mode, df, crawl_time):
    """将数据保存到数据库"""
    if df.empty:
        return
    
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    try:
        data_to_insert = []
        for _, row in df.iterrows():
            player_id = resolve_player_identity(row['name'], crawl_time, row['player_id'])
            if player_id is not None:
                data_to_insert.append((
                    player_id, row['player_id'], mode, row['rank'], row['name'], row['lv'], row['exp'],
                    row['acc'], row['combo'], row['pc'], crawl_time
                ))
        
        if data_to_insert:
            cursor.executemany('''
            INSERT OR IGNORE INTO player_rankings 
            (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', data_to_insert)
            
            db_manager.get_connection().commit()
            logger.info("模式 %d 的 %d 条数据已保存到数据库", mode, len(data_to_insert))
    except Exception as e:
        logger.error("保存模式 %d 数据到数据库失败: %s", mode, e)
        db_manager.get_connection().rollback()

def save_player_profile_to_database(player_data, crawl_time, player_identifier):
    """将玩家个人主页数据保存到数据库"""
    if not player_data:
        return False
    
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    try:
        data_to_insert = []
        for data in player_data:
            player_id = resolve_player_identity(data['name'], crawl_time, player_identifier)
            if player_id is not None:
                data_to_insert.append((
                    player_id, player_identifier, data['mode'], data['rank'], data['name'], data['lv'], data['exp'],
                    data['acc'], data['combo'], data['pc'], crawl_time
                ))
        
        if data_to_insert:
            cursor.executemany('''
            INSERT OR IGNORE INTO player_rankings 
            (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', data_to_insert)
            
            cursor.execute('''
            INSERT OR REPLACE INTO player_crawl_status 
            (player_identifier, last_crawled, crawl_count, success_count, last_error)
            VALUES (?, ?, COALESCE((SELECT crawl_count FROM player_crawl_status WHERE player_identifier = ?), 0) + 1, 
                   COALESCE((SELECT success_count FROM player_crawl_status WHERE player_identifier = ?), 0) + 1, NULL)
            ''', (player_identifier, crawl_time, player_identifier, player_identifier))
            
            db_manager.get_connection().commit()
            logger.info("玩家 %s 的 %d 条数据已保存到数据库", player_identifier, len(data_to_insert))
            return True
    except Exception as e:
        logger.error("保存玩家 %s 数据到数据库失败: %s", player_identifier, e)
        
        try:
            cursor.execute('''
            INSERT OR REPLACE INTO player_crawl_status 
            (player_identifier, last_crawled, crawl_count, success_count, last_error)
            VALUES (?, ?, COALESCE((SELECT crawl_count FROM player_crawl_status WHERE player_identifier = ?), 0) + 1, 
                   COALESCE((SELECT success_count FROM player_crawl_status WHERE player_identifier = ?), 0), ?)
            ''', (player_identifier, crawl_time, player_identifier, player_identifier, str(e)))
            db_manager.get_connection().commit()
        except:
            pass
            
        db_manager.get_connection().rollback()
        return False
    
    return False

def check_excel_file_integrity(filename):
    """检查Excel文件是否完整可用"""
    try:
        if not os.path.exists(filename):
            return False
            
        with open(filename, 'rb') as f:
            header = f.read(8)
            if header[:4] != b'PK\x03\x04':
                return False
                
        wb = load_workbook(filename, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        return bool(sheetnames)
    except Exception as e:
        logger.warning("Excel文件完整性检查失败: %s - %s", filename, e)
        return False

def repair_excel_file(filename):
    """尝试修复损坏的Excel文件"""
    try:
        backup_name = f"{filename}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        import shutil
        shutil.copy2(filename, backup_name)
        logger.info("已创建备份文件: %s", backup_name)
        
        try:
            xl = pd.ExcelFile(filename)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for sheet_name in xl.sheet_names:
                    df = pd.read_excel(filename, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info("成功修复Excel文件: %s", filename)
            return True
        except Exception as e:
            logger.error("使用pandas修复失败: %s", e)
            
            try:
                wb = load_workbook(filename)
                wb.save(filename)
                logger.info("使用openpyxl修复成功: %s", filename)
                return True
            except Exception as e2:
                logger.error("使用openpyxl修复也失败: %s", e2)
                return False
    except Exception as e:
        logger.error("修复Excel文件过程中发生错误: %s", e)
        return False

def import_mode_data(mode):
    """导入单个模式的数据"""
    global stop_requested
    
    thread_id = threading.get_ident()
    db_manager = DatabaseManager()
    conn = db_manager.get_connection(thread_id)
    
    conn.execute("PRAGMA synchronous = OFF")
    conn.execute("PRAGMA journal_mode = MEMORY")
    conn.execute("PRAGMA cache_size = 10000")
    conn.execute("PRAGMA temp_store = MEMORY")
    
    player_alias_cache = {}
    
    filename = get_excel_filename(mode)
    if not os.path.exists(filename):
        logger.warning("模式 %d 的Excel文件不存在: %s", mode, filename)
        return 0
    
    if not check_excel_file_integrity(filename):
        logger.warning("模式 %d 的Excel文件可能已损坏: %s", mode, filename)
        if repair_excel_file(filename):
            logger.info("文件修复成功，继续导入")
        else:
            logger.error("文件修复失败，跳过模式 %d", mode)
            return 0
    
    cursor = conn.cursor()
    cursor.execute(
        "SELECT last_import_time FROM import_metadata WHERE mode = ?",
        (mode,)
    )
    result = cursor.fetchone()
    last_import_time = result[0] if result else None
    
    try:
        xl = pd.ExcelFile(filename)
        sheet_names = xl.sheet_names
    except Exception as e:
        logger.error("打开Excel文件失败: %s", e)
        try:
            wb = load_workbook(filename)
            sheet_names = wb.sheetnames
        except Exception as e2:
            logger.error("两种方式都无法打开Excel文件: %s", e2)
            return 0
    
    sheet_names = [s for s in sheet_names if s.startswith(f"mode_{mode}_")]
    
    sheet_times = []
    for sheet_name in sheet_names:
        try:
            time_str = sheet_name.replace(f"mode_{mode}_", "")
            sheet_time = datetime.strptime(time_str, "%Y-%m-%d_%H-%M")
            sheet_times.append((sheet_name, sheet_time))
        except:
            continue
    
    sheet_times.sort(key=lambda x: x[1])
    
    if HAS_TQDM:
        mode_pbar = tqdm(
            sheet_times, 
            desc=f"模式 {mode}", 
            position=mode + 1, 
            leave=False,
            unit="表"
        )
    else:
        mode_pbar = sheet_times
        print(f"开始处理模式 {mode}，共 {len(sheet_times)} 个表...")
    
    imported_count = 0
    batch_size = 50
    batch_data = []
    
    for i, (sheet_name, sheet_time) in enumerate(mode_pbar):
        with stop_lock:
            if stop_requested:
                logger.info("模式 %d 导入被中断，已导入 %d 条数据", mode, imported_count)
                break
        
        if last_import_time and sheet_time <= datetime.strptime(last_import_time, "%Y-%m-%d %H:%M:%S"):
            if HAS_TQDM:
                mode_pbar.update(1)
            continue
            
        try:
            df = pd.read_excel(filename, sheet_name=sheet_name)
            if df.empty:
                if HAS_TQDM:
                    mode_pbar.update(1)
                continue
            
            for _, row in df.iterrows():
                name = row['name']
                
                if name in player_alias_cache:
                    player_id = player_alias_cache[name]
                else:
                    player_id = resolve_player_identity(name, sheet_time)
                    player_alias_cache[name] = player_id
                
                if player_id is not None:
                    batch_data.append((
                        player_id, None, mode, row['rank'], name, row['lv'], row['exp'],
                        row['acc'], row['combo'], row['pc'], sheet_time
                    ))
            
            if batch_data and (len(batch_data) >= 1000 or i == len(sheet_times) - 1):
                cursor.executemany('''
                INSERT OR IGNORE INTO player_rankings 
                (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', batch_data)
                
                imported_count += len(batch_data)
                batch_data = []
            
            if (i + 1) % batch_size == 0 or i == len(sheet_times) - 1:
                cursor.execute(
                    "UPDATE import_metadata SET last_import_time = ? WHERE mode = ?",
                    (sheet_time.strftime("%Y-%m-%d %H:%M:%S"), mode)
                )
                conn.commit()
            
            if HAS_TQDM:
                mode_pbar.set_postfix_str(f"已导入: {imported_count}")
            
        except Exception as e:
            logger.error("导入模式 %d 表 %s 时出错: %s", mode, sheet_name, e)
            conn.rollback()
            continue
    
    if batch_data:
        cursor.executemany('''
        INSERT OR IGNORE INTO player_rankings 
        (player_id, uid, mode, rank, name, lv, exp, acc, combo, pc, crawl_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', batch_data)
        imported_count += len(batch_data)
        conn.commit()
    
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.commit()
    
    db_manager.close_connection(thread_id)
    
    if HAS_TQDM:
        mode_pbar.close()
    
    return imported_count

def import_historical_data():
    """从Excel文件导入历史数据到数据库"""
    if HAS_TQDM:
        main_pbar = tqdm(total=len(MODES), desc="总体进度", position=0)
    else:
        print("开始导入历史数据...")
    
    for mode in MODES:
        try:
            result = import_mode_data(mode)
            if HAS_TQDM:
                main_pbar.update(1)
                main_pbar.set_postfix_str(f"模式 {mode} 完成: {result} 条记录")
            else:
                print(f"模式 {mode} 完成: {result} 条记录")
        except Exception as e:
            logger.error("模式 %d 导入失败: %s", mode, e)
            if HAS_TQDM:
                main_pbar.update(1)
                main_pbar.set_postfix_str(f"模式 {mode} 失败: {e}")
            else:
                print(f"模式 {mode} 失败: {e}")
    
    if HAS_TQDM:
        main_pbar.close()
    
    if not HAS_TQDM:
        print("历史数据导入完成")

def load_player_config():
    """加载玩家配置文件"""
    players = []
    
    if not os.path.exists(PLAYER_CONFIG_FILE):
        logger.info("玩家配置文件不存在，创建空文件: %s", PLAYER_CONFIG_FILE)
        with open(PLAYER_CONFIG_FILE, 'w', encoding='utf-8') as f:
            f.write("# 每行一个玩家ID（必须是数字）\n")
            f.write("# 例如:\n")
            f.write("# 923177\n")
            f.write("# 123456\n")
        return players
    
    try:
        with open(PLAYER_CONFIG_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    players.append(line)
        
        logger.info("从配置文件加载了 %d 个玩家", len(players))
        return players
    except Exception as e:
        logger.error("加载玩家配置文件失败: %s", e)
        return []

def add_players_to_queue(players):
    """将玩家添加到爬取队列"""
    for player in players:
        player_queue.put(player)
    logger.info("添加了 %d 个玩家到爬取队列", len(players))

def get_players_from_leaderboard(df_list):
    """从排行榜数据中提取玩家ID"""
    players = set()
    
    for df in df_list:
        if not df.empty and 'player_id' in df.columns:
            for player_id in df['player_id']:
                if player_id and player_id not in players:
                    players.add(player_id)
    
    return list(players)

def run_player_crawler():
    """运行玩家个人主页爬取器"""
    global player_crawl_in_progress, last_player_crawl_time
    
    with player_crawl_lock:
        if player_crawl_in_progress:
            logger.info("玩家爬取器已在运行，跳过")
            return
        
        player_crawl_in_progress = True
    
    try:
        logger.info("开始玩家个人主页爬取周期")
        last_player_crawl_time = datetime.now()
        
        session = requests.Session()
        session.cookies.update(COOKIES)
        session.headers.update(HEADERS)
        session.mount('https://', requests.adapters.HTTPAdapter(max_retries=3))
        
        total_players = player_queue.qsize()
        if total_players == 0:
            logger.info("爬取队列为空，跳过")
            return
        
        logger.info("开始爬取 %d 个玩家的个人主页数据", total_players)
        
        successful_crawls = 0
        failed_crawls = 0
        
        if HAS_TQDM:
            pbar = tqdm(total=total_players, desc="玩家主页爬取", unit="玩家")
        else:
            print(f"开始爬取 {total_players} 个玩家的个人主页数据...")
        
        while not player_queue.empty():
            with stop_lock:
                if stop_requested:
                    logger.info("玩家爬取被中断")
                    break
            
            try:
                player_identifier = player_queue.get_nowait()
            except queue.Empty:
                break
            
            try:
                player_data = crawl_player_profile(session, player_identifier)
                crawl_time = datetime.now()
                
                if player_data and save_player_profile_to_database(player_data, crawl_time, player_identifier):
                    successful_crawls += 1
                    if HAS_TQDM:
                        pbar.set_postfix_str(f"成功: {successful_crawls}, 失败: {failed_crawls}")
                else:
                    failed_crawls += 1
                    if HAS_TQDM:
                        pbar.set_postfix_str(f"成功: {successful_crawls}, 失败: {failed_crawls}")
                
                time.sleep(3)
                
            except Exception as e:
                logger.error("处理玩家 %s 时出错: %s", player_identifier, e)
                failed_crawls += 1
                if HAS_TQDM:
                    pbar.set_postfix_str(f"成功: {successful_crawls}, 失败: {failed_crawls}")
            finally:
                if HAS_TQDM:
                    pbar.update(1)
                player_queue.task_done()
        
        if HAS_TQDM:
            pbar.close()
        
        logger.info("玩家个人主页爬取完成: 成功 %d, 失败 %d", successful_crawls, failed_crawls)
        
    except Exception as e:
        logger.error("玩家爬取周期发生错误: %s", e)
    finally:
        with player_crawl_lock:
            player_crawl_in_progress = False

def start_player_crawler_thread():
    """启动玩家爬取器线程"""
    if player_queue.empty():
        logger.info("玩家队列为空，不启动爬取线程")
        return
    
    thread = Thread(target=run_player_crawler, daemon=True)
    thread.start()
    logger.info("玩家爬取器线程已启动")
    return thread

def git_check_updates():
    """检查远程Git仓库是否有更新"""
    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git更新检查")
            return False
            
        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)
        
        result = subprocess.run(["git", "remote", "-v"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("未配置Git远程仓库，跳过更新检查")
            os.chdir(original_cwd)
            return False
        
        result = subprocess.run(["git", "fetch", "origin"], capture_output=True, text=True)
        if result.returncode != 0:
            logger.warning("Git fetch失败: %s", result.stderr)
            os.chdir(original_cwd)
            return False
        
        result = subprocess.run(
            ["git", "diff", "--name-only", "HEAD", "origin/main", "--", "*.db", "*.xlsx"],
            capture_output=True, text=True
        )
        
        os.chdir(original_cwd)
        
        if result.stdout.strip():
            updated_files = result.stdout.strip().split('\n')
            logger.info("发现远程更新文件: %s", updated_files)
            return True
        else:
            logger.info("远程仓库没有.db或.xlsx文件的更新")
            return False
            
    except subprocess.CalledProcessError as e:
        logger.warning("Git检查更新失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git检查更新发生意外错误: %s", e)
        return False

def git_pull_data_files():
    """从远程Git仓库拉取文件"""
    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git拉取")
            return False
            
        print("\n" + "="*60)
        print("检测到远程仓库有更新！")
        print("更新文件包括: catch.xlsx, key.xlsx, malody_rankings.db 等数据文件")
        print("这些更新将覆盖您本地的数据文件。")
        print("="*60)
        print("您有10秒时间决定是否拉取更新：")
        print("  - 输入 'y' 或 'yes' 确认拉取")
        print("  - 输入 'n' 或 'no' 跳过拉取")
        print("  - 10秒内无响应将自动跳过")
        print("="*60)
        
        try:
            import select
            import sys
            
            print("请确认是否拉取更新 (y/n, 10秒超时): ", end='', flush=True)
            start_time = time.time()
            
            while time.time() - start_time < 10:
                if select.select([sys.stdin], [], [], 0.1)[0]:
                    user_input = sys.stdin.readline().strip().lower()
                    if user_input in ['y', 'yes']:
                        print("确认拉取更新...")
                        break
                    elif user_input in ['n', 'no']:
                        print("跳过拉取更新。")
                        return False
                    else:
                        print("无效输入，请输入 y 或 n: ", end='', flush=True)
                time.sleep(0.1)
            else:
                print("\n10秒超时，自动跳过拉取更新。")
                return False
                
        except ImportError:
            print("请确认是否拉取更新 (y/n, 10秒超时): ", end='', flush=True)
            user_input = input()
            if user_input.lower() in ['y', 'yes']:
                print("确认拉取更新...")
            else:
                print("跳过拉取更新。")
                return False
            
        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)
        
        DatabaseManager().close_connection()
        
        excel_files = [get_excel_filename(mode) for mode in MODES]
        files_to_pull = [f for f in excel_files if os.path.exists(f)]
        files_to_pull.append(DB_FILE)
        
        success = True
        for file in files_to_pull:
            result = subprocess.run(
                ["git", "checkout", "origin/main", "--", file],
                capture_output=True, text=True
            )
            if result.returncode == 0:
                logger.info("已拉取文件: %s", file)
            else:
                logger.warning("拉取文件 %s 失败: %s", file, result.stderr)
                success = False
        
        os.chdir(original_cwd)
        return success
    except subprocess.CalledProcessError as e:
        logger.warning("Git拉取文件失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git拉取文件发生意外错误: %s", e)
        return False

def git_add_commit_push(has_changes=True):
    """自动添加、提交和推送Git更改"""
    if not has_changes:
        logger.info("所有模式均无数据变化，跳过Git推送")
        return True
        
    try:
        if not os.path.exists(os.path.join(GIT_REPO_PATH, '.git')):
            logger.info("当前目录不是Git仓库，跳过Git推送")
            return True
            
        original_cwd = os.getcwd()
        os.chdir(GIT_REPO_PATH)
        
        result = subprocess.run(["git", "remote", "-v"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("未配置Git远程仓库，跳过推送")
            os.chdir(original_cwd)
            return True
        
        excel_files = [get_excel_filename(mode) for mode in MODES]
        files_to_add = [f for f in excel_files if os.path.exists(f)]
        files_to_add.append(DB_FILE)
        
        for file in files_to_add:
            result = subprocess.run(["git", "add", file], capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning("添加文件 %s 失败: %s", file, result.stderr)
        
        result = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True)
        if not result.stdout.strip():
            logger.info("没有文件更改，跳过Git提交")
            os.chdir(original_cwd)
            return True
        
        result = subprocess.run(["git", "commit", "-m", GIT_COMMIT_MESSAGE], capture_output=True, text=True)
        if result.returncode == 0:
            logger.info("Git提交成功: %s", GIT_COMMIT_MESSAGE)
        else:
            logger.warning("Git提交失败: %s", result.stderr)
            os.chdir(original_cwd)
            return False
        
        result = subprocess.run(["git", "push"], capture_output=True, text=True)
        if result.returncode == 0:
            logger.info("Git推送成功")
            success = True
        else:
            logger.warning("Git推送失败: %s", result.stderr)
            success = False
        
        os.chdir(original_cwd)
        return success
    except subprocess.CalledProcessError as e:
        logger.warning("Git操作失败: %s", e)
        return False
    except Exception as e:
        logger.warning("Git操作发生意外错误: %s", e)
        return False

def check_data_changed(mode, df):
    """检查数据是否发生变化"""
    filename = get_excel_filename(mode)
    sheet_name = f"mode_{mode}"
    
    if not os.path.exists(filename):
        return True
    
    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return True
        
        sub_sheets = [s for s in wb.sheetnames if s.startswith(f"{sheet_name}_")]
        if not sub_sheets:
            return True
            
        latest_sheet = None
        latest_time = None
        for s in sub_sheets:
            try:
                dt_str = s.replace(f"{sheet_name}_", "")
                dt = datetime.strptime(dt_str, "%Y-%m-%d_%H-%M")
                if latest_time is None or dt > latest_time:
                    latest_time = dt
                    latest_sheet = s
            except:
                continue
        
        if latest_sheet:
            df_prev = pd.read_excel(filename, sheet_name=latest_sheet)
            
            if not df_prev.empty and df_prev.equals(df):
                return False
                
    except Exception as e:
        logger.error("检查数据变化时出错: %s", e)
        return True
        
    return True

def run_crawler_cycle(crawl_players=False, save_excel=False):
    """运行爬取周期
    
    Args:
        crawl_players: 是否爬取玩家主页数据
        save_excel: 是否保存数据到Excel文件
    """
    try:
        if git_check_updates():
            logger.info("检测到远程仓库有更新，正在拉取数据文件...")
            if git_pull_data_files():
                logger.info("数据文件拉取完成，重新初始化数据库...")
                DatabaseManager().close_connection()
                init_database()
            else:
                logger.warning("数据文件拉取失败，继续使用本地数据")
        else:
            logger.info("未检测到远程更新或Git不可用，继续使用本地数据")
    except Exception as e:
        logger.warning("Git更新检查失败，继续使用本地数据: %s", e)
    
    session = requests.Session()
    session.cookies.update(COOKIES)
    session.headers.update(HEADERS)
    
    session.mount('https://', requests.adapters.HTTPAdapter(max_retries=3))

    start_time = datetime.now()
    logger.info("=" * 50)
    logger.info("开始爬取周期: %s", start_time)

    has_changes = False
    all_dfs = []
    
    for mode in MODES:
        try:
            with stop_lock:
                if stop_requested:
                    logger.info("爬取被中断")
                    break
            
            logger.info("处理模式: %d", mode)
            df = crawl_mode_player(session, mode)
            all_dfs.append(df)
            
            if df.empty:
                logger.warning("模式 %d 获取数据为空，跳过", mode)
                continue
                
            if save_excel and not check_data_changed(mode, df):
                logger.info("模式 %d 数据未变化，跳过保存", mode)
                continue
                
            crawl_time = datetime.now()
            
            # 只有在启用Excel保存时才保存到Excel
            if save_excel:
                save_data_to_excel(mode, df, crawl_time)
                
            # 始终保存到数据库
            save_to_database(mode, df, crawl_time)
            has_changes = True
            
            time.sleep(3)
        except Exception as e:
            logger.exception("处理模式 %d 时发生错误", mode)
    
    if crawl_players and all_dfs:
        leaderboard_players = get_players_from_leaderboard(all_dfs)
        if leaderboard_players:
            add_players_to_queue(leaderboard_players)
            logger.info("从排行榜添加了 %d 个玩家到爬取队列", len(leaderboard_players))
        
        start_player_crawler_thread()
    
    try:
        git_add_commit_push(has_changes)
    except Exception as e:
        logger.warning("Git推送失败，但数据已保存到本地: %s", e)
    
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    logger.info("爬取周期完成, 用时: %.2f秒", duration)
    logger.info("=" * 50)

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='Malody排行榜爬虫')
    parser.add_argument('--all', action='store_true', 
                       help='爬取排行榜和玩家主页数据')
    parser.add_argument('--leaderboard-only', action='store_true', 
                       help='只爬取排行榜数据（默认）')
    parser.add_argument('--players-only', action='store_true', 
                       help='只爬取玩家主页数据')
    parser.add_argument('--migrate-db', action='store_true', 
                       help='执行数据库迁移（添加uid字段）')
    parser.add_argument('--import-only', action='store_true', 
                       help='只导入历史数据')
    parser.add_argument('--once', action='store_true', 
                       help='运行一次爬取周期后退出')
    parser.add_argument('--save-excel', action='store_true',
                       help='保存数据到Excel文件（默认不保存）')
    
    return parser.parse_args()

def run_players_only():
    """只运行玩家主页爬取"""
    init_database()
    config_players = load_player_config()
    if config_players:
        add_players_to_queue(config_players)
        run_player_crawler()
    else:
        logger.info("没有配置玩家，跳过爬取")
    DatabaseManager().close_connection()

def main():
    args = parse_arguments()
    
    if args.migrate_db:
        migrate_database()
        return
    
    init_database()
    
    config_players = load_player_config()
    if config_players:
        add_players_to_queue(config_players)
        logger.info("从配置文件加载了 %d 个玩家", len(config_players))
    
    db_manager = DatabaseManager()
    cursor = db_manager.get_connection().cursor()
    
    cursor.execute("SELECT COUNT(*) FROM player_rankings")
    count = cursor.fetchone()[0]
    
    if count == 0:
        logger.info("开始导入历史数据...")
        import_historical_data()
        logger.info("历史数据导入完成")
    else:
        logger.info("数据库中已有 %d 条记录，跳过历史数据导入", count)
    
    if args.import_only:
        DatabaseManager().close_connection()
        return
    elif args.players_only:
        run_players_only()
        return
    elif args.once:
        run_crawler_cycle(crawl_players=args.all, save_excel=args.save_excel)
        DatabaseManager().close_connection()
        return
    else:
        try:
            while True:
                with stop_lock:
                    if stop_requested:
                        logger.info("程序被终止")
                        break
                
                try:
                    run_crawler_cycle(crawl_players=args.all, save_excel=args.save_excel)
                except Exception as e:
                    logger.exception("主循环发生未处理异常")
                
                logger.info("等待30分钟后重启...")
                
                for i in range(30):
                    with stop_lock:
                        if stop_requested:
                            logger.info("程序被终止")
                            break
                    
                    time.sleep(60)
                    gc.collect()
        finally:
            DatabaseManager().close_connection()

if __name__ == "__main__":
    main()