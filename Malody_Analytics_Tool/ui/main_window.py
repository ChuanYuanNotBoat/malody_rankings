# ui/main_window.py
import os
import logging
import pandas as pd
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
  QMainWindow, QAction, QActionGroup, QMessageBox,
  QFileDialog, QStatusBar, QLabel, QVBoxLayout, QWidget, QApplication,
  QTabWidget, QDateEdit, QHBoxLayout, QPushButton, QTableWidget,
  QTableWidgetItem, QHeaderView, QComboBox, QProgressBar
)
from PyQt5.QtCore import QSettings, QEvent, QTranslator, Qt, QDateTime, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QIcon
from core.analytics import analyze_mode_data, MODE_FILES, get_latest_sheet_data
from core.history_analyzer import get_player_history, get_all_players_growth
from widgets.chart_widget import ChartWidget
from widgets.history_chart import HistoryChartWidget
import concurrent.futures
import psutil
from core.history_analyzer import calculate_player_growth


logger = logging.getLogger(__name__)

# 设置最大内存使用
MAX_MEMORY = psutil.virtual_memory().available * 0.7  # 使用70%可用内存


# ================= 后台线程类 =================
class ParallelAnalysisThread(QThread):
  finished = pyqtSignal(dict)  # 所有模式的分析结果
  progress = pyqtSignal(int, int)  # 当前进度, 总任务数
  error = pyqtSignal(int, str)  # 模式, 错误信息

  def __init__(self, folder_path):
    super().__init__()
    self.folder_path = folder_path
    self.results = {}
    self.cancel_requested = False

  def run(self):
    try:
      mode_files = list(MODE_FILES.items())
      total = len(mode_files)

      # 使用线程池并行处理
      with concurrent.futures.ThreadPoolExecutor(
        max_workers=min(4, os.cpu_count())
      ) as executor:
        futures = {}
        for mode, filename in mode_files:
          if self.cancel_requested:
            break

          file_path = os.path.join(self.folder_path, filename)
          futures[executor.submit(self.analyze_mode, mode, file_path)] = mode

        completed = 0
        for future in concurrent.futures.as_completed(futures):
          if self.cancel_requested:
            break

          mode = futures[future]
          try:
            result = future.result()
            if result:
              self.results[mode] = result
          except Exception as e:
            self.error.emit(mode, str(e))

          completed += 1
          self.progress.emit(completed, total)

      if not self.cancel_requested:
        self.finished.emit(self.results)

    except Exception as e:
      logger.error(f"Parallel analysis failed: {str(e)}")

  def analyze_mode(self, mode, file_path):
    """分析单个模式"""
    if not os.path.exists(file_path):
      return None

    try:
      df = get_latest_sheet_data(file_path)
      if df.empty:
        return None
      return analyze_mode_data(df, mode)
    except Exception as e:
      logger.error(f"Error analyzing mode {mode}: {str(e)}")
      raise e

  def cancel(self):
    """取消分析任务"""
    self.cancel_requested = True


class HistoryThread(QThread):
  finished = pyqtSignal(str, list)  # 玩家名称, 历史数据
  error = pyqtSignal(str)

  def __init__(self, folder_path, player_name, start_date, end_date, max_points=500):
    super().__init__()
    self.folder_path = folder_path
    self.player_name = player_name
    self.start_date = start_date
    self.end_date = end_date
    self.max_points = max_points

  def run(self):
    try:
      history = get_player_history(self.folder_path, self.player_name)

      # 过滤日期范围
      filtered_history = [
        h for h in history
        if h['date'].date() >= self.start_date and h['date'].date() <= self.end_date
      ]

      # 如果数据点太多，进行采样
      if len(filtered_history) > self.max_points:
        step = max(1, len(filtered_history) // self.max_points)
        filtered_history = filtered_history[::step]

      self.finished.emit(self.player_name, filtered_history)
    except Exception as e:
      self.error.emit(str(e))


class GrowthThread(QThread):
  finished = pyqtSignal(dict)  # 成长数据
  error = pyqtSignal(str)
  progress = pyqtSignal(int, int)  # 当前进度, 总任务数

  def __init__(self, folder_path, start_date, end_date):
    super().__init__()
    self.folder_path = folder_path
    self.start_date = start_date
    self.end_date = end_date
    self.cancel_requested = False

  def run(self):
    try:
      # 首先获取所有玩家列表
      all_players = set()
      for mode, filename in MODE_FILES.items():
        file_path = os.path.join(self.folder_path, filename)
        if not os.path.exists(file_path):
          continue

        try:
          # 使用只读模式
          wb = load_workbook(file_path, read_only=True)
          for sheet_name in wb.sheetnames:
            if not sheet_name.startswith(f"mode_{mode}_"):
              continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # 找到玩家名称所在的列
            try:
              name_col_idx = headers.index("name")
            except ValueError:
              continue

            # 收集玩家名称
            for row in ws.iter_rows(min_row=2):
              if name_col_idx < len(row) and row[name_col_idx].value:
                all_players.add(row[name_col_idx].value)

          wb.close()
        except Exception as e:
          logger.error(f"Error loading players from {file_path}: {str(e)}")
          continue

      player_list = list(all_players)
      total_players = len(player_list)
      player_growth = {}

      # 处理每个玩家
      for idx, player in enumerate(player_list):
        if self.cancel_requested:
          break

        try:
          # 确保正确传递文件夹路径
          history = get_player_history(self.folder_path, player)

          # 计算成长数据 - 使用正确的函数
          growth = calculate_player_growth(history, self.start_date, self.end_date)
          if growth:
            player_growth[player] = growth
        except Exception as e:
          logger.error(f"Error calculating growth for {player}: {str(e)}")
        # 每处理10个玩家更新一次进度
        if idx % 10 == 0:
          self.progress.emit(idx + 1, total_players)

      if not self.cancel_requested:
        self.finished.emit(player_growth)
    except Exception as e:
      self.error.emit(str(e))

  def cancel(self):
    """取消任务"""
    self.cancel_requested = True


# ================= 主窗口类 =================
class MainWindow(QMainWindow):
  def __init__(self, translator):
    super().__init__()
    self.translator = translator
    self.settings = QSettings("MalodyAnalytics", "MalodyAnalyticsTool")
    self.folder_path = ""
    self.player_history = []
    self.player_growth_data = {}
    self.init_ui()
    self.load_language()

    # 内存监控
    self.memory_timer = QTimer(self)
    self.memory_timer.timeout.connect(self.monitor_memory)
    self.memory_timer.start(5000)  # 每5秒检查一次内存

  def monitor_memory(self):
    """监控内存使用情况"""
    mem = psutil.virtual_memory()
    if mem.percent > 90:
      logger.warning(f"内存使用过高: {mem.percent}%")
      self.status_label.setText(self.tr("警告: 内存使用过高! 请关闭其他应用"))

  def init_ui(self):
    # 设置窗口标题和图标
    self.setWindowTitle(self.tr("Malody Analytics Tool"))
    self.setWindowIcon(QIcon(":/icons/app_icon.png"))
    self.setGeometry(100, 100, 1200, 800)

    # 创建菜单栏
    self.create_menus()

    # 创建状态栏
    self.status_bar = QStatusBar()
    self.setStatusBar(self.status_bar)
    self.status_label = QLabel(self.tr("Ready"))
    self.status_bar.addWidget(self.status_label)

    # 添加进度条
    self.progress_bar = QProgressBar()
    self.progress_bar.setMaximum(100)
    self.progress_bar.setVisible(False)
    self.status_bar.addPermanentWidget(self.progress_bar)

    # 在状态栏添加取消按钮
    self.cancel_button = QPushButton(self.tr("取消操作"))
    self.cancel_button.clicked.connect(self.cancel_operations)
    self.cancel_button.setVisible(False)
    self.status_bar.addPermanentWidget(self.cancel_button)

    # 创建主内容区域
    self.main_widget = QTabWidget()
    self.setCentralWidget(self.main_widget)

    # 模式分析标签页
    self.mode_tab = QWidget()
    self.mode_layout = QVBoxLayout(self.mode_tab)

    # 创建图表控件
    self.chart_widget = ChartWidget()
    self.mode_layout.addWidget(self.chart_widget)
    self.main_widget.addTab(self.mode_tab, self.tr("Mode Analysis"))

    # 玩家历史分析标签页
    self.history_tab = QWidget()
    self.history_layout = QVBoxLayout(self.history_tab)

    # 玩家选择控件
    player_control_layout = QHBoxLayout()
    self.history_layout.addLayout(player_control_layout)

    player_control_layout.addWidget(QLabel(self.tr("Select Player:")))
    self.player_combo = QComboBox()
    self.player_combo.currentIndexChanged.connect(self.on_player_changed)
    player_control_layout.addWidget(self.player_combo)

    # 日期范围控件
    date_control_layout = QHBoxLayout()
    self.history_layout.addLayout(date_control_layout)

    date_control_layout.addWidget(QLabel(self.tr("Date Range:")))
    self.start_date_edit = QDateEdit()
    self.start_date_edit.setCalendarPopup(True)
    self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
    self.start_date_edit.setDate(QDateTime.currentDateTime().addMonths(-1).date())
    date_control_layout.addWidget(self.start_date_edit)

    date_control_layout.addWidget(QLabel(self.tr("to")))
    self.end_date_edit = QDateEdit()
    self.end_date_edit.setCalendarPopup(True)
    self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
    self.end_date_edit.setDate(QDateTime.currentDateTime().date())
    date_control_layout.addWidget(self.end_date_edit)

    self.refresh_btn = QPushButton(self.tr("Refresh History"))
    self.refresh_btn.clicked.connect(self.refresh_player_history)
    date_control_layout.addWidget(self.refresh_btn)

    date_control_layout.addStretch()

    # 历史图表
    self.history_chart = HistoryChartWidget()
    self.history_layout.addWidget(self.history_chart)
    self.main_widget.addTab(self.history_tab, self.tr("Player History"))

    # 玩家成长分析标签页
    self.growth_tab = QWidget()
    self.growth_layout = QVBoxLayout(self.growth_tab)

    # 成长分析控件
    growth_control_layout = QHBoxLayout()
    self.growth_layout.addLayout(growth_control_layout)

    growth_control_layout.addWidget(QLabel(self.tr("Sort By:")))
    self.sort_combo = QComboBox()
    self.sort_combo.addItem(self.tr("Rank Change (Best First)"), "rank_change")
    self.sort_combo.addItem(self.tr("Level Change (Best First)"), "lv_change")
    self.sort_combo.addItem(self.tr("Experience Change (Best First)"), "exp_change")
    self.sort_combo.addItem(self.tr("Play Count Change (Best First)"), "pc_change")
    self.sort_combo.addItem(self.tr("Daily EXP Growth (Best First)"), "daily_exp_growth")
    self.sort_combo.addItem(self.tr("Player Name (A-Z)"), "name")
    growth_control_layout.addWidget(self.sort_combo)

    self.refresh_growth_btn = QPushButton(self.tr("Refresh Growth Data"))
    self.refresh_growth_btn.clicked.connect(self.refresh_growth_data)
    growth_control_layout.addWidget(self.refresh_growth_btn)

    growth_control_layout.addStretch()

    # 成长数据表格
    self.growth_table = QTableWidget()
    self.growth_table.setColumnCount(8)
    self.growth_table.setHorizontalHeaderLabels([
      self.tr("Player"),
      self.tr("Rank Change"),
      self.tr("Level Change"),
      self.tr("EXP Change"),
      self.tr("Play Count Change"),
      self.tr("Daily EXP Growth"),
      self.tr("Period"),
      self.tr("Days")
    ])
    self.growth_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    self.growth_table.setSortingEnabled(True)
    self.growth_table.setSelectionBehavior(QTableWidget.SelectRows)
    self.growth_table.setEditTriggers(QTableWidget.NoEditTriggers)
    self.growth_layout.addWidget(self.growth_table)
    self.main_widget.addTab(self.growth_tab, self.tr("Player Growth"))

    # 连接模式选择信号
    self.chart_widget.mode_combo.currentIndexChanged.connect(self.on_mode_selected)

  def on_mode_selected(self):
    """当用户选择模式时触发"""
    mode = self.chart_widget.mode_combo.currentData()

    # "All Modes" 不需要单独分析
    if mode is None:
      self.chart_widget.on_mode_changed()
      return

    # 如果数据已存在，直接显示
    if mode in self.chart_widget.mode_data:
      self.chart_widget.on_mode_changed()
      return

    # 启动分析线程
    self.analyze_mode(mode)

  def analyze_mode(self, mode):
    """分析指定模式的数据"""
    if not self.folder_path:
      return

    file_path = os.path.join(self.folder_path, MODE_FILES.get(mode, f"mode{mode}.xlsx"))

    # 如果已有分析线程在运行，跳过
    if mode in self.analysis_threads:
      return

    self.status_label.setText(self.tr("Analyzing mode {}...").format(mode))

    # 创建并启动分析线程
    thread = ModeAnalysisThread(mode, file_path)
    thread.finished.connect(self.on_mode_analysis_finished)
    thread.error.connect(self.on_mode_analysis_error)
    thread.start()

    self.analysis_threads[mode] = thread

  def on_mode_analysis_finished(self, mode, data):
    """模式分析完成"""
    # 从线程字典中移除
    if mode in self.analysis_threads:
      del self.analysis_threads[mode]

    # 更新图表数据
    self.chart_widget.update_mode_data(mode, data)
    self.status_label.setText(self.tr("Mode {} analysis completed").format(mode))

  def on_mode_analysis_error(self, mode, error_msg):
    """模式分析出错"""
    # 从线程字典中移除
    if mode in self.analysis_threads:
      del self.analysis_threads[mode]

    QMessageBox.critical(self, self.tr("Error"),
                         self.tr("Failed to analyze mode {}: {}").format(mode, error_msg))
    self.status_label.setText(self.tr("Ready"))

  def create_menus(self):
    """创建菜单栏"""
    # 文件菜单
    file_menu = self.menuBar().addMenu(self.tr("&File"))

    self.open_action = QAction(QIcon(":/icons/open.png"), self.tr("&Open Folder"), self)
    self.open_action.setShortcut("Ctrl+O")
    self.open_action.triggered.connect(self.open_folder)
    file_menu.addAction(self.open_action)

    self.save_action = QAction(QIcon(":/icons/save.png"), self.tr("&Save Report"), self)
    self.save_action.setShortcut("Ctrl+S")
    self.save_action.triggered.connect(self.save_report)
    file_menu.addAction(self.save_action)

    file_menu.addSeparator()

    self.exit_action = QAction(self.tr("E&xit"), self)
    self.exit_action.setShortcut("Ctrl+Q")
    self.exit_action.triggered.connect(self.close)
    file_menu.addAction(self.exit_action)

    # 语言菜单
    lang_menu = self.menuBar().addMenu(self.tr("&Language"))

    self.chinese_action = QAction(self.tr("中文"), self)
    self.chinese_action.triggered.connect(lambda: self.set_language('zh_CN'))
    self.chinese_action.setCheckable(True)
    lang_menu.addAction(self.chinese_action)

    self.english_action = QAction("English", self)
    self.english_action.triggered.connect(lambda: self.set_language('en'))
    self.english_action.setCheckable(True)
    lang_menu.addAction(self.english_action)

    # 创建互斥选项组
    self.language_group = QActionGroup(self)
    self.language_group.addAction(self.chinese_action)
    self.language_group.addAction(self.english_action)
    self.language_group.setExclusive(True)

    # 帮助菜单
    help_menu = self.menuBar().addMenu(self.tr("&Help"))

    self.about_action = QAction(self.tr("&About"), self)
    self.about_action.triggered.connect(self.show_about)
    help_menu.addAction(self.about_action)

  def load_language(self):
    """加载语言设置"""
    lang = self.settings.value("language", "")
    if not lang:
      lang = get_system_language()
      logger.debug(f"No saved language, using system: {lang}")
    else:
      logger.debug(f"Loaded saved language: {lang}")

    self.set_language(lang, initial_load=True)

  def set_language(self, lang_code, initial_load=False):
    """设置应用程序语言"""
    # 如果语言没有变化，直接返回
    current_lang = self.settings.value("language", "")
    if current_lang == lang_code and not initial_load:
      return

    # 更新UI选择状态
    if lang_code == 'zh_CN':
      self.chinese_action.setChecked(True)
      self.english_action.setChecked(False)
    else:
      self.chinese_action.setChecked(False)
      self.english_action.setChecked(True)

    # 保存语言设置
    self.settings.setValue("language", lang_code)
    self.settings.sync()  # 确保立即写入磁盘

    logger.debug(f"Language set to: {lang_code}, saved to settings")

    # 重新加载翻译器
    QApplication.removeTranslator(self.translator)
    self.translator = QTranslator()

    if lang_code == 'zh_CN':
      # 尝试多种方式加载中文翻译
      loaded = False

      # 方式1：从资源文件加载
      if self.translator.load(":/translations/malody_zh_CN.qm"):
        QApplication.installTranslator(self.translator)
        loaded = True
        logger.info("Loaded Chinese translation from resource")

      # 方式2：从文件系统加载
      if not loaded:
        try:
          script_dir = os.path.dirname(os.path.abspath(__file__))
          qm_path = os.path.join(script_dir, "..", "translations", "malody_zh_CN.qm")
          if os.path.exists(qm_path) and self.translator.load(qm_path):
            QApplication.installTranslator(self.translator)
            loaded = True
            logger.info(f"Loaded Chinese translation from file: {qm_path}")
          else:
            logger.error(f"Failed to load translation file: {qm_path}")
        except Exception as e:
          logger.error(f"Error loading translation file: {str(e)}")

      if not loaded:
        logger.warning("Failed to load Chinese translation")

    # 触发语言更改事件
    self.retranslate_ui()

    # 只有在语言实际变更时才发送事件
    if current_lang != lang_code:
      event = QEvent(QEvent.LanguageChange)
      QApplication.sendEvent(self, event)

    # 更新图表控件
    self.chart_widget.retranslate_ui()

    # 添加标签页标题翻译
    self.main_widget.setTabText(0, self.tr("Mode Analysis"))
    self.main_widget.setTabText(1, self.tr("Player History"))
    self.main_widget.setTabText(2, self.tr("Player Growth"))

    # 提示用户重启应用
    if not initial_load:
      QMessageBox.information(
        self,
        self.tr("Language Changed"),
        self.tr("The application needs to restart for the language change to take full effect.")
      )

  def retranslate_ui(self):
    """重新翻译所有UI文本"""
    logger.debug("Retranslating UI elements")

    # 更新所有UI文本
    self.setWindowTitle(self.tr("Malody Analytics Tool"))
    self.open_action.setText(self.tr("&Open Folder"))
    self.save_action.setText(self.tr("&Save Report"))
    self.exit_action.setText(self.tr("E&xit"))
    self.chinese_action.setText(self.tr("中文"))
    self.english_action.setText(self.tr("English"))
    self.about_action.setText(self.tr("&About"))
    self.status_label.setText(self.tr("Ready"))

    # 添加成长表格标题翻译
    self.growth_table.setHorizontalHeaderLabels([
      self.tr("Player"),
      self.tr("Rank Change"),
      self.tr("Level Change"),
      self.tr("EXP Change"),
      self.tr("Play Count Change"),
      self.tr("Daily EXP Growth"),
      self.tr("Period"),
      self.tr("Days")
    ])
    self.refresh_btn.setText(self.tr("Refresh History"))
    self.refresh_growth_btn.setText(self.tr("Refresh Growth Data"))
    self.sort_combo.setItemText(0, self.tr("Rank Change (Best First)"))
    self.sort_combo.setItemText(1, self.tr("Level Change (Best First)"))
    self.sort_combo.setItemText(2, self.tr("Experience Change (Best First)"))
    self.sort_combo.setItemText(3, self.tr("Play Count Change (Best First)"))
    self.sort_combo.setItemText(4, self.tr("Daily EXP Growth (Best First)"))
    self.sort_combo.setItemText(5, self.tr("Player Name (A-Z)"))

    # 更新菜单标题
    file_menu = self.menuBar().actions()[0].menu()
    file_menu.setTitle(self.tr("&File"))

    lang_menu = self.menuBar().actions()[1].menu()
    lang_menu.setTitle(self.tr("&Language"))

    help_menu = self.menuBar().actions()[2].menu()
    help_menu.setTitle(self.tr("&Help"))

  def open_folder(self):
    """打开文件夹对话框"""
    folder_path = QFileDialog.getExistingDirectory(
      self,
      self.tr("Select Malody Data Folder"),
      ""
    )

    if folder_path:
      self.folder_path = folder_path
      self.status_label.setText(self.tr("Processing folder..."))
      self.progress_bar.setVisible(True)
      self.progress_bar.setRange(0, len(MODE_FILES))
      self.open_action.setEnabled(False)
      self.cancel_button.setVisible(True)

      # 启动并行分析线程
      self.analysis_thread = ParallelAnalysisThread(folder_path)
      self.analysis_thread.finished.connect(self.on_analysis_completed)
      self.analysis_thread.progress.connect(self.update_analysis_progress)
      self.analysis_thread.error.connect(self.on_analysis_error)
      self.analysis_thread.start()

  def update_analysis_progress(self, completed, total):
    """更新分析进度"""
    self.progress_bar.setValue(completed)
    self.status_label.setText(
      self.tr("Processing: {}/{} modes").format(completed, total)
    )

  def on_analysis_completed(self, results):
    """分析完成处理"""
    self.progress_bar.setVisible(False)
    self.open_action.setEnabled(True)
    self.cancel_button.setVisible(False)

    self.chart_widget.mode_data = results
    self.chart_widget.on_mode_changed()

    # 加载玩家列表
    self.load_player_list()

    # 加载成长数据
    self.refresh_growth_data()

    self.status_label.setText(self.tr("Folder loaded successfully"))

  def on_analysis_error(self, mode, error_msg):
    """分析出错"""
    QMessageBox.critical(self, self.tr("Error"),
                         self.tr("Failed to analyze mode {}: {}").format(mode, error_msg))
    self.status_label.setText(self.tr("Ready"))

  def load_player_list(self):
    """加载玩家列表"""
    if not self.folder_path:
      return

    try:
      all_players = set()
      for mode, filename in MODE_FILES.items():
        file_path = os.path.join(self.folder_path, filename)
        if not os.path.exists(file_path):
          continue

        try:
          # 使用只读模式
          wb = load_workbook(file_path, read_only=True)
          for sheet_name in wb.sheetnames:
            if not sheet_name.startswith(f"mode_{mode}_"):
              continue

            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # 找到玩家名称所在的列
            try:
              name_col_idx = headers.index("name")
            except ValueError:
              continue

            # 收集玩家名称
            for row in ws.iter_rows(min_row=2):
              if name_col_idx < len(row) and row[name_col_idx].value:
                all_players.add(row[name_col_idx].value)

          wb.close()
        except Exception as e:
          logger.error(f"Error loading players from {file_path}: {str(e)}")
          continue

      # 更新玩家下拉框
      self.player_combo.clear()
      if all_players:
        sorted_players = sorted(all_players)
        self.player_combo.addItems(sorted_players)
        self.history_chart.set_player_list(sorted_players)
        self.player_combo.setCurrentIndex(0)

    except Exception as e:
      logger.error(f"Error loading player list: {str(e)}")

  def save_report(self):
    """保存分析报告"""
    QMessageBox.information(self, self.tr("Save"), self.tr("Report saving functionality will be implemented soon."))

  def show_about(self):
    """显示关于对话框"""
    about_text = self.tr(
      "<h2>Malody Analytics Tool</h2>"
      "<p>Version 1.3.0</p>"
      "<p>A tool for analyzing Malody rhythm game ranking data.</p>"
      "<p>© 2023 Malody Analytics Team</p>"
    )
    QMessageBox.about(self, self.tr("About"), about_text)

  def changeEvent(self, event):
    """处理语言更改事件"""
    if event.type() == QEvent.LanguageChange:
      logger.debug("LanguageChange event received")
      self.retranslate_ui()
    super().changeEvent(event)

  def on_player_changed(self, index):
    """玩家选择改变时触发"""
    if index >= 0 and self.folder_path:
      player_name = self.player_combo.currentText()
      self.refresh_player_history()

  def refresh_player_history(self):
    """刷新玩家历史数据"""
    if not self.folder_path or self.player_combo.count() == 0:
      return

    player_name = self.player_combo.currentText()
    start_date = self.start_date_edit.date().toPyDate()
    end_date = self.end_date_edit.date().toPyDate()

    self.status_label.setText(self.tr("Loading history for {}...").format(player_name))
    self.refresh_btn.setEnabled(False)
    self.cancel_button.setVisible(True)

    # 创建并启动历史数据线程
    self.history_thread = HistoryThread(
      self.folder_path,
      player_name,
      start_date,
      end_date,
      max_points=500  # 限制最多500个数据点
    )
    self.history_thread.finished.connect(self.on_history_finished)
    self.history_thread.error.connect(self.on_history_error)
    self.history_thread.start()

  def on_history_finished(self, player_name, history):
    self.refresh_btn.setEnabled(True)
    self.cancel_button.setVisible(False)
    self.history_chart.set_history_data(player_name, history)
    self.status_label.setText(self.tr("History loaded for {}").format(player_name))

  def on_history_error(self, error_msg):
    self.refresh_btn.setEnabled(True)
    self.cancel_button.setVisible(False)
    QMessageBox.critical(self, self.tr("Error"), self.tr("Failed to load history: {}").format(error_msg))
    self.status_label.setText(self.tr("Ready"))

  def refresh_growth_data(self):
    """刷新玩家成长数据"""
    if not self.folder_path:
      return

    start_date = self.start_date_edit.date().toPyDate()
    end_date = self.end_date_edit.date().toPyDate()

    self.status_label.setText(self.tr("Calculating player growth..."))
    self.refresh_growth_btn.setEnabled(False)
    self.progress_bar.setVisible(True)
    self.progress_bar.setRange(0, 100)  # 不确定进度
    self.cancel_button.setVisible(True)

    # 创建并启动成长数据线程
    self.growth_thread = GrowthThread(self.folder_path, start_date, end_date)
    self.growth_thread.finished.connect(self.on_growth_finished)
    self.growth_thread.error.connect(self.on_growth_error)
    self.growth_thread.progress.connect(self.update_growth_progress)
    self.growth_thread.start()

  def update_growth_progress(self, current, total):
    """更新成长数据进度"""
    self.progress_bar.setValue(int(current / total * 100))
    self.status_label.setText(
      self.tr("Processing players: {}/{}").format(current, total)
    )

  def on_growth_finished(self, growth_data):
    self.refresh_growth_btn.setEnabled(True)
    self.progress_bar.setVisible(False)
    self.cancel_button.setVisible(False)
    self.player_growth_data = growth_data
    self.update_growth_table()
    self.status_label.setText(self.tr("Growth data updated"))

  def on_growth_error(self, error_msg):
    self.refresh_growth_btn.setEnabled(True)
    self.progress_bar.setVisible(False)
    self.cancel_button.setVisible(False)
    QMessageBox.critical(self, self.tr("Error"), self.tr("Failed to calculate growth: {}").format(error_msg))
    self.status_label.setText(self.tr("Ready"))

  def cancel_operations(self):
    """取消所有后台操作"""
    if hasattr(self, 'analysis_thread') and self.analysis_thread.isRunning():
      self.analysis_thread.cancel()
      self.analysis_thread.quit()
      self.analysis_thread.wait()
      self.status_label.setText(self.tr("Operation cancelled"))
      self.progress_bar.setVisible(False)
      self.open_action.setEnabled(True)

    if hasattr(self, 'history_thread') and self.history_thread.isRunning():
      self.history_thread.terminate()
      self.history_thread.wait()
      self.status_label.setText(self.tr("Operation cancelled"))
      self.refresh_btn.setEnabled(True)

    if hasattr(self, 'growth_thread') and self.growth_thread.isRunning():
      self.growth_thread.cancel()
      self.growth_thread.quit()
      self.growth_thread.wait()
      self.status_label.setText(self.tr("Operation cancelled"))
      self.progress_bar.setVisible(False)
      self.refresh_growth_btn.setEnabled(True)

    self.cancel_button.setVisible(False)

  def update_growth_table(self):
    """更新成长数据表格"""
    if not self.player_growth_data:
      self.growth_table.setRowCount(0)
      return

    # 获取排序方式
    sort_field = self.sort_combo.currentData()

    # 准备数据
    data = []
    for player, growth in self.player_growth_data.items():
      # 跳过没有有效数据的变化
      if growth.get('rank_change') is None and growth.get('lv_change') is None:
        continue

      data.append({
        'player': player,
        'rank_change': growth.get('rank_change', 0),
        'lv_change': growth.get('lv_change', 0),
        'exp_change': growth.get('exp_change', 0),
        'pc_change': growth.get('pc_change', 0),
        'daily_exp_growth': growth.get('daily_exp_growth', 0),
        'period': f"{growth['start_date'].strftime('%Y-%m-%d')} to {growth['end_date'].strftime('%Y-%m-%d')}",
        'days': growth.get('days', 0)
      })

    # 排序数据
    if sort_field == "name":
      data.sort(key=lambda x: x['player'])
    else:
      # 对于数值字段，降序排列（最佳变化在前）
      data.sort(key=lambda x: x.get(sort_field, 0) or 0, reverse=True)

    # 更新表格
    self.growth_table.setRowCount(len(data))

    for row, item in enumerate(data):
      self.growth_table.setItem(row, 0, QTableWidgetItem(item['player']))
      self.growth_table.setItem(row, 1, QTableWidgetItem(str(item['rank_change'])))
      self.growth_table.setItem(row, 2, QTableWidgetItem(str(item['lv_change'])))
      self.growth_table.setItem(row, 3, QTableWidgetItem(str(item['exp_change'])))
      self.growth_table.setItem(row, 4, QTableWidgetItem(str(item['pc_change'])))
      self.growth_table.setItem(row, 5, QTableWidgetItem(str(item['daily_exp_growth'])))
      self.growth_table.setItem(row, 6, QTableWidgetItem(item['period']))
      self.growth_table.setItem(row, 7, QTableWidgetItem(str(item['days'])))
