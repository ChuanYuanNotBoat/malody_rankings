import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import logging
import numpy as np
import pyarrow as pa

# 添加logger定义
logger = logging.getLogger(__name__)

# 模式文件命名规则
MODE_FILES = {
  0: "key.xlsx",
  3: "catch.xlsx"
}
# 其他模式
for i in range(1, 10):
  if i != 3:  # 跳过3，因为已经定义了
    MODE_FILES[i] = f"mode{i}.xlsx"


def get_latest_sheet_data(file_path):
  """获取指定Excel文件中最新工作表的数据"""
  try:
    # 获取所有工作表名
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames

    # 过滤并排序工作表（按日期降序）
    valid_sheets = []
    for name in sheet_names:
      match = re.match(r"mode_(\d+)_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2})", name)
      if match:
        mode, date_str, time_str = match.groups()
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        time_obj = datetime.strptime(time_str, "%H-%M")
        # 组合日期和时间
        combined = datetime.combine(date_obj.date(), time_obj.time())
        valid_sheets.append((name, combined))

    if not valid_sheets:
      logger.warning(f"No valid sheets found in {file_path}")
      return pd.DataFrame()

    # 按日期降序排序
    valid_sheets.sort(key=lambda x: x[1], reverse=True)
    latest_sheet = valid_sheets[0][0]
    logger.debug(f"Found {len(valid_sheets)} valid sheets, using latest: {latest_sheet}")

    # 读取数据 - 移除 chunksize 参数
    df = pd.read_excel(file_path, sheet_name=latest_sheet)

    # 确保有数据
    if df.empty:
      logger.warning(f"Latest sheet '{latest_sheet}' in {file_path} is empty")

    return df

  except Exception as e:
    logger.error(f"Error reading {file_path}: {str(e)}")
    return pd.DataFrame()


def analyze_mode_data(df, mode):
  """分析一个模式的数据，返回结果字典"""
  if df.empty:
    logger.warning(f"No data found for mode {mode}")
    return {}

  # 基本统计
  total_players = len(df)
  avg_accuracy = df['acc'].mean()
  avg_play_count = df['pc'].mean()
  top_player = df.iloc[0].to_dict() if total_players > 0 else {}

  # 准确率分布
  bins = [0, 70, 80, 90, 100]
  labels = ['<70%', '70-79%', '80-89%', '90-100%']
  accuracy_distribution = pd.cut(df['acc'], bins=bins, labels=labels, right=False).value_counts().to_dict()

  # 等级分布
  level_distribution = df['lv'].value_counts().sort_index().to_dict()

  return {
    "mode": mode,
    "total_players": total_players,
    "avg_accuracy": round(avg_accuracy, 2),
    "avg_play_count": round(avg_play_count),
    "top_player": top_player,
    "accuracy_distribution": accuracy_distribution,
    "level_distribution": level_distribution
  }


def analyze_malody_folder(folder_path):
  """分析文件夹中的所有Malody模式文件"""
  results = {}
  logger.info(f"Analyzing Malody folder: {folder_path}")

  for mode, filename in MODE_FILES.items():
    file_path = os.path.join(folder_path, filename)

    if not os.path.exists(file_path):
      logger.warning(f"File not found: {file_path}")
      continue

    try:
      # 读取最新数据
      df = get_latest_sheet_data(file_path)

      if df.empty:
        logger.warning(f"No valid data found in {file_path}")
        continue

      # 分析模式数据
      mode_results = analyze_mode_data(df, mode)
      results[mode] = mode_results
      logger.info(f"Analyzed mode {mode}: {mode_results['total_players']} players")

    except Exception as e:
      logger.error(f"Error analyzing mode {mode}: {str(e)}")

  return results
