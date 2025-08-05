import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import logging
import re
import threading
from concurrent.futures import ThreadPoolExecutor
import psutil

# 从analytics导入MODE_FILES
from core.analytics import MODE_FILES

logger = logging.getLogger(__name__)

# 预编译正则表达式
SHEET_PATTERN = re.compile(r"mode_(\d+)_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2})")

# 设置最大内存使用
MAX_MEMORY = psutil.virtual_memory().available * 0.7  # 使用70%可用内存


def parse_sheet_date(sheet_name):
  """从工作表名解析日期"""
  match = SHEET_PATTERN.match(sheet_name)
  if match:
    mode_num = int(match.group(1))
    date_str = match.group(2)
    try:
      return datetime.strptime(date_str, "%Y-%m-%d_%H-%M"), mode_num
    except ValueError:
      return None, None
  return None, None


class PlayerHistoryCache:
  """玩家历史数据缓存"""

  def __init__(self, folder_path):
    self.folder_path = folder_path
    self.cache = {}
    self.lock = threading.Lock()

  def get_history(self, player_name):
    """获取玩家历史，使用缓存"""
    with self.lock:
      if player_name in self.cache:
        return self.cache[player_name]

      history = self._load_player_history(player_name)
      self.cache[player_name] = history
      return history

  def preload_players(self, player_names):
    """预加载多个玩家的历史数据"""
    with ThreadPoolExecutor(max_workers=min(4, os.cpu_count())) as executor:
      futures = {executor.submit(self.get_history, name): name for name in player_names}

      for future in futures:
        try:
          future.result()  # 确保加载完成
        except Exception as e:
          logger.error(f"Error preloading player history: {str(e)}")

  def _load_player_history(self, player_name):
    """实际加载玩家历史数据"""
    history = {}

    # 遍历所有模式文件
    for mode, filename in MODE_FILES.items():
      file_path = os.path.join(self.folder_path, filename)

      if not os.path.exists(file_path):
        continue

      try:
        # 使用只读模式
        wb = load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
          dt, mode_num = parse_sheet_date(sheet_name)
          if dt is None or mode_num != mode:
            continue

          ws = wb[sheet_name]
          headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

          # 找到玩家名称所在的列
          try:
            name_col_idx = headers.index("name")
          except ValueError:
            continue

          # 查找玩家数据
          player_found = False
          for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
            if row[name_col_idx].value == player_name:
              player_data = {}
              for idx, cell in enumerate(row):
                if idx < len(headers):
                  player_data[headers[idx]] = cell.value
              player_data["date"] = dt
              player_data["mode"] = mode
              history[dt] = player_data
              player_found = True
              break

          if not player_found:
            # 记录玩家在该时间点不存在
            history[dt] = {
              "player": player_name,
              "date": dt,
              "mode": mode,
              "status": "not_found"
            }

        wb.close()
      except Exception as e:
        logger.error(f"Error processing {file_path}: {str(e)}")

    # 转换为按日期排序的列表
    sorted_history = sorted(history.values(), key=lambda x: x["date"])
    return sorted_history


# core/history_analyzer.py

def get_player_history(folder_path, player_name):
  """
  获取指定玩家的历史数据
  :param folder_path: 数据文件夹路径
  :param player_name: 玩家名称
  :return: 历史数据列表，按日期排序
  """
  history = []
  player_files = {}  # 存储玩家出现的文件

  # 首先收集所有包含该玩家的文件
  for mode, filename in MODE_FILES.items():
    file_path = os.path.join(folder_path, filename)
    if not os.path.exists(file_path):
      continue

    try:
      wb = load_workbook(file_path, read_only=True)
      for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(f"mode_{mode}_"):
          continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        try:
          name_col_idx = headers.index("name")
        except ValueError:
          continue

        # 检查玩家是否在该工作表中
        player_found = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
          if name_col_idx < len(row) and row[name_col_idx].value == player_name:
            player_found = True
            break

        if player_found:
          # 提取日期
          match = re.match(r"mode_(\d+)_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2})", sheet_name)
          if match:
            date_str = match.group(2)
            try:
              date_obj = datetime.strptime(date_str, "%Y-%m-%d")
              player_files[(file_path, sheet_name)] = date_obj
            except ValueError:
              continue

      wb.close()
    except Exception as e:
      logger.error(f"Error scanning {file_path} for player {player_name}: {str(e)}")

  # 按日期排序文件
  sorted_files = sorted(player_files.items(), key=lambda x: x[1])

  # 处理每个文件
  for (file_path, sheet_name), date_obj in sorted_files:
    try:
      df = pd.read_excel(file_path, sheet_name=sheet_name)

      # 找到玩家数据
      player_data = df[df['name'] == player_name].iloc[0].to_dict()

      # 添加日期
      player_data['date'] = date_obj

      history.append(player_data)
    except Exception as e:
      logger.error(f"Error processing {file_path} sheet {sheet_name}: {str(e)}")

  # 按日期排序
  history.sort(key=lambda x: x['date'])
  return history


def calculate_player_growth(history, start_date=None, end_date=None):
  """计算玩家在指定日期范围内的成长"""
  if not history:
    return {}

  # 过滤日期范围
  if start_date:
    history = [h for h in history if h['date'].date() >= start_date]
  if end_date:
    history = [h for h in history if h['date'].date() <= end_date]

  if not history:
    return {}

  # 计算增量
  first = history[0]
  last = history[-1]

  return {
    'player_name': last.get('name', 'Unknown'),
    'rank_change': last['rank'] - first['rank'] if 'rank' in last and 'rank' in first else None,
    'lv_change': last['lv'] - first['lv'] if 'lv' in last and 'lv' in first else None,
    'exp_change': last['exp'] - first['exp'] if 'exp' in last and 'exp' in first else None,
    'pc_change': last['pc'] - first['pc'] if 'pc' in last and 'pc' in first else None,
    'days': (last['date'] - first['date']).days,
    'daily_exp_growth': round((last['exp'] - first['exp']) / (last['date'] - first['date']).days, 2)
    if 'exp' in last and 'exp' in first and (last['date'] - first['date']).days > 0 else None,
    'start_date': first['date'],
    'end_date': last['date']
  }


def get_all_players_growth(folder_path, start_date=None, end_date=None):
  """获取所有玩家在指定日期范围内的成长数据"""
  player_growth = {}

  # 首先获取所有玩家列表
  all_players = set()
  for mode, filename in MODE_FILES.items():
    file_path = os.path.join(folder_path, filename)
    if not os.path.exists(file_path):
      continue

    try:
      # 使用只读模式
      wb = load_workbook(file_path, read_only=True)
      for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(f"mode_{mode}_"):
          continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # 找到玩家名称所在的列
        try:
          name_col_idx = headers.index("name")
        except ValueError:
          continue

        # 收集玩家名称
        for row in ws.iter_rows(min_row=2):
          if name_col_idx < len(row) and row[name_col_idx].value:
            all_players.add(row[name_col_idx].value)

      wb.close()
    except Exception as e:
      logger.error(f"Error loading players from {file_path}: {str(e)}")
      continue

  # 计算每个玩家的成长
  for player in all_players:
    history = get_player_history(folder_path, player)
    growth = calculate_player_growth(history, start_date, end_date)
    if growth:
      player_growth[player] = growth

  return player_growth
